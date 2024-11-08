# Copyright (c) <2024> <Yusuf Bunyamin Kahveci>

# Permission is hereby granted, free of charge, to any person obtaining a copy of this software and
# associated documentation files (the "Software"), to deal in the Software without restriction,
# including without limitation the rights to use, copy, modify, merge, publish, distribute,
# sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all copies or
# substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT
# NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT
# OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# get jlc bom and altium bom as arguments
if len(sys.argv) != 3:
    print("Usage: python jlc_bomcheck.py <altium_bom_path> <jlc_bom_path>")
    sys.exit(1)

altium_bom_path = sys.argv[1]
jlc_bom_path = sys.argv[2]

# Load the XLS file
# file_path = 'BOM_DDS RX 1-ch_REV4_B_2024-10-03.xlsx'
altium_xls_data = pd.read_excel(altium_bom_path)

# Delete first 4 rows to remove the header
altium_xls_data = altium_xls_data.iloc[4:]

# first row is the header
altium_xls_data.columns = altium_xls_data.iloc[0]
altium_xls_data = altium_xls_data[1:]

# remove every column but designator and manufacturer part number and supplier part number
altium_xls_data = altium_xls_data[
    ["Designator", "Manufacturer Part Number 1", "Supplier Part Number 1"]
]

# create a data structure to hold the data as
# {designator: [manufacturer part number, supplier part number]}
# repeat the process for each designator number
altium_bom_data = {}
for index, row in altium_xls_data.iterrows():
    designator_cell = row["Designator"]
    manufacturer_part_number = row["Manufacturer Part Number 1"]
    supplier_part_number = row["Supplier Part Number 1"]
    # parse designator column to get real designator
    # each designator is separated by a comma and a space
    designator_list = designator_cell.split(", ")
    for designator in designator_list:
        if designator in altium_bom_data:
            print("Designator already exists")
            continue
        altium_bom_data[designator] = []
        altium_bom_data[designator].append(
            [manufacturer_part_number, supplier_part_number]
        )


# --------------------------------------------------------------------------------------------
# Load the BOM from JLCPCB
# jlc_bom_path = 'bom JLC.xls'
jlc_xls_data = pd.read_excel(jlc_bom_path)

# remove all columns except for top designator, bottom designator and JLCPCB Part #
jlc_xls_data = jlc_xls_data[["topDesignator", "bottomDesignator", "JLCPCB Part #"]]

jlc_bom_data = {}
for index, row in jlc_xls_data.iterrows():
    topdesignator_cell = row["topDesignator"]
    bottomdesignator_cell = row["bottomDesignator"]
    jlc_part_number = row["JLCPCB Part #"]

    # parse designator column to get real designator, each designator is separated by a comma
    if pd.notna(topdesignator_cell):
        topdesignator_list = topdesignator_cell.split(",")
        for designator in topdesignator_list:
            if designator in jlc_bom_data:
                print("Designator already exists")
                continue
            jlc_bom_data[designator] = []
            jlc_bom_data[designator].append([jlc_part_number])
    if pd.notna(bottomdesignator_cell):
        bottomdesignator_list = bottomdesignator_cell.split(",")
        for designator in bottomdesignator_list:
            if designator in jlc_bom_data:
                print("Designator already exists")
                continue
            jlc_bom_data[designator] = []
            jlc_bom_data[designator].append([jlc_part_number])


# Compare the same designators and check if the supplier part number and JLCPCB Part # match
for designator in altium_bom_data:
    if designator in jlc_bom_data:
        if altium_bom_data[designator][0][1] == jlc_bom_data[designator][0][0]:
            # print("Designator: " + designator + " matches")
            pass
        else:
            print(
                "Designator: "
                + designator
                + " does not match. Supplier Part Number: "
                + str(altium_bom_data[designator][0][1])
                + " JLC Part Number: "
                + str(jlc_bom_data[designator][0][0])
            )
    else:
        print("Designator: " + designator + " not found in JLC BOM")

# create a new excel file with the results. it will have the designator, the supplier part number
# and the JLC part number, the result of the comparison and the reason,
results = []
for designator in altium_bom_data:
    if designator in jlc_bom_data:
        if altium_bom_data[designator][0][1] == jlc_bom_data[designator][0][0]:
            results.append(
                [
                    designator,
                    altium_bom_data[designator][0][0],
                    altium_bom_data[designator][0][1],
                    jlc_bom_data[designator][0][0],
                    "Match",
                ]
            )
        elif pd.isna(altium_bom_data[designator][0][1]):  # empty
            results.append(
                [
                    designator,
                    altium_bom_data[designator][0][0],
                    altium_bom_data[designator][0][1],
                    jlc_bom_data[designator][0][0],
                    "No Match",
                ]
            )
        else:
            results.append(
                [
                    designator,
                    altium_bom_data[designator][0][0],
                    altium_bom_data[designator][0][1],
                    jlc_bom_data[designator][0][0],
                    "Double Check",
                ]
            )
    else:
        results.append(
            [
                designator,
                altium_bom_data[designator][0][0],
                altium_bom_data[designator][0][1],
                "",
                "Not Found",
            ]
        )

results_df = pd.DataFrame(
    results,
    columns=[
        "Designator",
        "Manufacturer PN",
        "Supplier Part Number",
        "JLC Part Number",
        "Result",
    ],
)
results_df.to_excel("results.xlsx", index=False)

# Color the background of the result cells
# red for no match, green for match, yellow for not found and orange for double check
# We need to use openpyxl for this

wb = load_workbook("results.xlsx")
ws = wb.active

# Define the fill colors
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Apply the fill colors based on the Result column
for row in range(2, len(results) + 2):
    cell = ws.cell(row=row, column=5)  # Result column is the 5th column
    if cell.value == "No Match":
        cell.fill = red_fill
    elif cell.value == "Match":
        cell.fill = green_fill
    elif cell.value == "Not Found":
        cell.fill = yellow_fill
    elif cell.value == "Double Check":
        cell.fill = orange_fill

#Adjust the column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save("results.xlsx")
print("Results saved to results.xlsx")
